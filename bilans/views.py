from django.shortcuts import render
from django.views.generic import (TemplateView, ListView, CreateView, DetailView, FormView)
from . models import Canva, Bilan, Autre, Tresorerie, Production, Checkbox
from django.contrib import messages
from django.views.generic.detail import SingleObjectMixin
from django.urls import reverse
from .forms import BilansFormset, AutresFormset, TresoreriesFormset, ProductionsFormset, CheckboxsFormset
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from .decorators import allowed_users
from django.utils import timezone
import xlwt
from xlutils.copy import copy # http://pypi.python.org/pypi/xlutils
from xlrd import open_workbook # http://pypi.python.org/pypi/xlrd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter


def export_bilans_xls(request, pk):    
   
    bilan_queryset = Bilan.objects.filter(canva_id=pk)
    
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    for c in Canva.objects.filter(id=pk):
        s = c.site
        d = c.mois      
        a = c.année

    response['Content-Disposition'] = 'attachment; filename=Agregats Mensuels - {site} - {mois} - {année}.xlsx'.format(site=s, mois=d, année=a)

    workbook = Workbook()
    
    worksheet = workbook.active
    worksheet.title = 'TB SCF Ecofie'
    
    
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['N°', 'Agregat', 'SCF', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 7
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for b in bilan_queryset:
        row_num += 1
        n=n+1
        
        if b.mois2 is None:
            b.mois2=0
        if b.mois1 is None:
            b.mois1=0
            row = [n,
                b.agregat, 
                b.SCF, 
                b.mois1,
                b.mois2,
                b.mois2-b.mois1, 
                100, 
                b.finmois1, 
                b.finmois2, 
                b.ecart2, 
                b.evolution2,]
    
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
                cell.alignment = Alignment('center')  
                 
            
                if cell.value is None:
                    cell.value=0
        
        else:
            
            row = [n,
                b.agregat, 
                b.SCF, 
                b.mois1, 
                b.mois2,
                b.mois2-b.mois1, 
                (b.mois2-b.mois1)/b.mois1*100, 
                b.finmois1, 
                b.finmois2, 
                b.ecart2, 
                b.evolution2,]
    
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
                cell.alignment = Alignment('center')    
            
                if cell.value is None:
                    cell.value=0
                    
    
    
    
    
    top = worksheet.cell (row=1, column=2)
    top.value = "TABLEAU DE BORD MENSUEL"
    top.font = Font(name='Calibri', bold=True, size=20, color='000000ff')

    b = 6
    for i in range(14,34):
        ag = worksheet.cell (row=i, column=1)
        ag.value = b
        b = b+1
    
    
    
    
    
    

    
    
    
    
    top2 = worksheet.cell (row=1, column=8)
    for c in Canva.objects.filter(id=pk):
        top2.value = c.mois
        top2.font = Font(name='Calibri', bold=True, size=20)
    
    top3 = worksheet.cell (row=1, column=10)
    for c in Canva.objects.filter(id=pk):
        top3.value = c.année
        top3.font = Font(name='Calibri', bold=True, size=20)
    
    top4 = worksheet.cell (row=3, column=2)
    top4.value = "Groupe"
    top4.border = Border(bottom=Side(border_style='thick'), left=Side(border_style='thick'), top=Side(border_style='thick'), right=Side(border_style='thick'))
    top4.font = Font(name='Calibri', bold=True, size=13)
    top4.alignment = Alignment(horizontal='center')
    
    top5 = worksheet.cell (row=4, column=2)
    top5.value = "EPE"
    top5.alignment = Alignment(horizontal='center')
    top5.border = Border(left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    top5.font = Font(name='Calibri', bold=True, size=13)
    
    top6 = worksheet.cell (row=3, column=5)
    top6.value = "IMETAL"
    top6.border = Border(bottom=Side(border_style='thick'), top=Side(border_style='thick'))
    top6.font = Font(name='Calibri', bold=True, size=13, color='000000ff')
    
    top7 = worksheet.cell (row=4, column=5)
    top7.value = "ANABIB"
    top7.border = Border(bottom=Side(border_style='thick'))
    top7.font = Font(name='Calibri', bold=True, size=13, color='000000ff')
    
    top8 = worksheet.cell (row=5, column=10)
    top8.value = "En millions de DA"
    top8.font = Font(name='Calibri', bold=True, size=13, color='00ff0000')
    
    vide1 = worksheet.cell (row=3, column=3)
    vide1.border = Border(bottom=Side(border_style='thick'), top=Side(border_style='thick'))
    vide2 = worksheet.cell (row=3, column=4)
    vide2.border = Border(bottom=Side(border_style='thick'), top=Side(border_style='thick'))
    vide3 = worksheet.cell (row=3, column=6)
    vide3.border = Border(bottom=Side(border_style='thick'), top=Side(border_style='thick'), right=Side(border_style='thick'))
    vide4 = worksheet.cell (row=4, column=3)
    vide4.border = Border(bottom=Side(border_style='thick'))
    vide5 = worksheet.cell (row=4, column=4)
    vide5.border = Border(bottom=Side(border_style='thick'))
    vide6 = worksheet.cell (row=4, column=6)
    vide6.border = Border(bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    
    
    
    
    
    
    
    
    column_letter = get_column_letter(1)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 5
    
    
    column_letter = get_column_letter(2)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 45
    
    column_letter = get_column_letter(3)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 13
    
    for i in range(4,12):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 17
    
    column_letter = get_column_letter(8)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 20
    
    column_letter = get_column_letter(9)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 20
    
    
    
    for i in range(8,34):
        ag = worksheet.cell (row=i, column=2)
        ag.alignment = Alignment('left')
    
    
    
    ind1 = worksheet.cell (row=20, column=2)
    ind1.alignment = Alignment(indent=2)
    
    ind2 = worksheet.cell (row=21, column=2)
    ind2.alignment = Alignment(indent=5)
    
    ind3 = worksheet.cell (row=22, column=2)
    ind3.alignment = Alignment(indent=5)
    
    worksheet.merge_cells('A13:A14')
    merged1 = worksheet['A13']
    merged1.value = "6"
    merged1.alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet.merge_cells('A6:A7')
    merged2 = worksheet['A6']
    merged2b = worksheet['A7']
    merged2.value = 'N°'
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('B6:B7')
    merged3 = worksheet['B6']
    merged3b = worksheet['B7']
    merged3.value = 'AGREGATS'
    merged3.alignment = Alignment(horizontal="center", vertical="center")
    merged3.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged3b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged3.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged3.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('C6:C7')
    merged4 = worksheet['C6']
    merged4b = worksheet['C7']
    merged4.value = 'Comptes SCF'
    merged4.alignment = Alignment(horizontal="center", vertical="center")
    merged4.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged4b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged4.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged4.font = Font(name='Calibri', bold=True, size=11)
    
    worksheet.merge_cells('D6:E6')
    merged5 = worksheet['D6']
    merged5b = worksheet['E6']
    merged5.value = 'Mois'
    merged5.alignment = Alignment(horizontal="center", vertical="center")
    merged5.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged5b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged5.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged5.font = Font(name='Calibri', bold=True, size=13)
    
    
    worksheet.merge_cells('F6:G6')
    merged6 = worksheet['F6']
    merged6b = worksheet['G6']
    merged6.value = 'Evolution'
    merged6.alignment = Alignment(horizontal="center", vertical="center")
    merged6.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged6b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged6.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged6.font = Font(name='Calibri', bold=True, size=13, color='000000ff')
    
    worksheet.merge_cells('H6:I6')
    merged7 = worksheet['H6']
    merged7b = worksheet['I6']
    merged7.value = 'Cumul'
    merged7.alignment = Alignment(horizontal="center", vertical="center")
    merged7.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged7b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged7.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged7.font = Font(name='Calibri', bold=True, size=13)
    
    
    
    worksheet.merge_cells('J6:K6')
    merged8 = worksheet['J6']
    merged8b = worksheet['K6']
    merged8.value = 'Evolution'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=13, color='000000ff')
    
    head1 = worksheet.cell (row=7, column=6)
    head1.font = Font(color='000000ff')
    
    head2 = worksheet.cell (row=7, column=7)
    head2.font = Font(color='000000ff')
    
    head3 = worksheet.cell (row=7, column=10)
    head3.font = Font(color='000000ff')
    
    head4 = worksheet.cell (row=7, column=11)
    head4.font = Font(color='000000ff')

    
    for i in range(1,12):
        line1 = worksheet.cell (row=13, column=i)
        line1.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'))
        line1.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        line2 = worksheet.cell (row=14, column=i)
        line2.border =  Border(bottom=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), top=Side(border_style='medium'))
        line2.fill = PatternFill(start_color="00ffcc99", end_color=("00ffcc99"), fill_type='solid')
        line3 = worksheet.cell (row=18, column=i)
        line3.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line3.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        line4 = worksheet.cell (row=25, column=i)
        line4.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line4.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        line5 = worksheet.cell (row=26, column=i)
        line5.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line5.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        line6 = worksheet.cell (row=29, column=i)
        line6.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line6.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        line7 = worksheet.cell (row=33, column=i)
        line7.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line7.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    
    
    
    
    test1 = worksheet.cell (row=8, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=4)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=4)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=4)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=4)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=5)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=5)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=5)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=5)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=5)
    sum1.value = int(value1+value2+value3+value4+value5) 
    
    test1 = worksheet.cell (row=8, column=6)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=6)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=6)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=6)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=6)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    
    
    
    
    test1 = worksheet.cell (row=13, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=13, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=13, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    
    
    
    
    
    test1 = worksheet.cell (row=8, column=8)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=8)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=8)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=8)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=8)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=9)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=9)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=9)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=9)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=9)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=10)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=10)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=10)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=10)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=10)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=11)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=11)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=11)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=11)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=11)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=11)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    
    
    test1 = worksheet.cell (row=13, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=4)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=4)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=4)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=13, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=5)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=5)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=5)
    sum1.value = int(value1+value2+value3+value4)
    sum1.number_format = '### ### ###'
    
    test1 = worksheet.cell (row=13, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=6)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=6)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=6)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=18, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=18, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=18, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    test1 = worksheet.cell (row=13, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=8)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=8)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=8)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=13, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=9)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=9)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=9)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=13, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=10)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=10)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=10)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=19, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=4)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=4)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=5)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=5)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=6)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=6)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=25, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=25, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=25, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    test1 = worksheet.cell (row=19, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=8)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=8)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=9)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=9)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=10)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=10)
    sum1.value = int(value1+value2+value3)
    
    
    cell = worksheet.cell (row=8, column=5)
    cell.number_format = '### ### ### ###'
    
    
    
    test1 = worksheet.cell (row=18, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=4)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=4)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=5)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=6)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=6)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=8)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=8)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=9)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=9)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=10)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=10)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=26, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=26, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    test1 = worksheet.cell (row=26, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=4)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=4)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=5)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=5)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=6)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=6)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=8)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=8)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=9)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=9)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=10)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=10)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=29, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=29, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=29, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    
    
    
    
    
    
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=6)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
        
    for i in range(8,34):   
        cell = worksheet.cell (row=i, column=7)
        cell.number_format = '0.00'
    
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=8)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=9)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
           
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=10)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    
    
    
    
    
    
    
    
    
    
    
    
    
    autre_queryset = Autre.objects.filter(canva_id=pk)
    
    for c in Canva.objects.filter(id=pk):
        columns = ['N°', 'Autres Agregats', 'Comptes SCF', 'a la fin Mois N-1', 'a la fin Mois N', 'Ecart en Valeur', '% evolution', 'Observations:']

    row_num = 36
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for a in autre_queryset:
        row_num += 1
        n=n+1
        
        if a.finmoisn is None:
            a.finmoisn=0
        if a.finmoisn1 is None:
            a.finmoisn1=0
            row = [n, a.autreag, a.SCF, a.finmoisn1, a.finmoisn, a.finmoisn-a.finmoisn1, 100,a.observation]
    
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
                cell.alignment = Alignment('center')    
            
                if cell.value is None:
                    cell.value=0
        
        else:
            row = [n, a.autreag, a.SCF, a.finmoisn1, a.finmoisn, a.finmoisn-a.finmoisn1, (a.finmoisn-a.finmoisn1)/a.finmoisn1*100,a.observation]
    
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
                cell.alignment = Alignment('center')    
            
                if cell.value is None:
                    cell.value=0
        
        
    
    
    
            
    
    
    
    
    for i in range(37,51):
        ag = worksheet.cell (row=i, column=2)
        ag.alignment = Alignment('left')
    
    ind1 = worksheet.cell (row=38, column=2)
    ind1.alignment = Alignment(indent=2)
    ind1.font = Font(color='00008000')
    
    ind1 = worksheet.cell (row=38, column=3)
    ind1.font = Font(color='00008000')
    
    ind2 = worksheet.cell (row=39, column=2)
    ind2.alignment = Alignment(indent=5)
    
    ind3 = worksheet.cell (row=40, column=2)
    ind3.alignment = Alignment(indent=8)
    ind3.font = Font(color='00ff0000')
    
    ind3 = worksheet.cell (row=40, column=3)
    ind3.font = Font(color='00ff0000')
    
    ind4 = worksheet.cell (row=42, column=2)
    ind4.alignment = Alignment(indent=5)
    
    ind5 = worksheet.cell (row=44, column=2)
    ind5.alignment = Alignment(indent=2)
    
    ind6 = worksheet.cell (row=45, column=2)
    ind6.alignment = Alignment(indent=5)
    
    ind7 = worksheet.cell (row=46, column=2)
    ind7.alignment = Alignment(indent=5)
    
    ind8 = worksheet.cell (row=48, column=2)
    ind8.alignment = Alignment(indent=5)
    
    ind9 = worksheet.cell (row=50, column=2)
    ind9.alignment = Alignment(indent=5)
    
    worksheet.merge_cells('B35:B36')
    merged8 = worksheet['B35']
    merged8b = worksheet['B36']
    merged8.value = 'Autres Agrégats'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('C35:C36')
    merged8 = worksheet['C35']
    merged8b = worksheet['C36']
    merged8.value = 'Comptes SCF'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=11)
    
    worksheet.merge_cells('D35:E35')
    merged8 = worksheet['D35']
    merged8b = worksheet['E35']
    merged8.value = 'Cumul'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('F35:G35')
    merged8 = worksheet['F35']
    merged8b = worksheet['G35']
    merged8.value = 'Evolution'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=13, color='000000ff')
    
    worksheet.merge_cells('H35:K36')
    merged8 = worksheet['H35']
    merged8b = worksheet['K36']
    merged8.value = 'Observations :'
    merged8.alignment = Alignment(horizontal="left", vertical="top")
    merged8.border =  Border(left=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099cc00", end_color=("0099cc00"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A35:A36')
    merged8 = worksheet['A35']
    merged8b = worksheet['A36']
    merged8.value = ''
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A37:A40')
    merged8 = worksheet['A37']
    merged8b = worksheet['A40']
    merged8.value = '25'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A41:A42')
    merged8 = worksheet['A41']
    merged8b = worksheet['A42']
    merged8.value = '26'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A43:A46')
    merged8 = worksheet['A43']
    merged8b = worksheet['A46']
    merged8.value = '27'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A47:A48')
    merged8 = worksheet['A47']
    merged8b = worksheet['A48']
    merged8.value = '28'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A49:A50')
    merged8 = worksheet['A49']
    merged8b = worksheet['A50']
    merged8.value = '29'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    
    
    
    for i in range(8, 12):
        for n in range(37, 51):
            ob = worksheet.cell (row=n, column=i)
            ob.alignment = Alignment('left')
            ob.alignment = Alignment(horizontal="left", vertical="top")
            ob.fill = PatternFill(start_color="0099cc00", end_color=("0099cc00"), fill_type='solid')
            ob.font = Font(name='Calibri', bold=True, size=11)
            ob.border =  Border(left=Side(border_style='none'))
            if ob.value==0:
                ob.value=" "
        
    
    head1 = worksheet.cell (row=36, column=6)
    head1.font = Font(color='000000ff')
    
    head1 = worksheet.cell (row=36, column=7)
    head1.font = Font(color='000000ff')
    
    top = worksheet.cell (row=52, column=1)
    top.value = "* Veuillez renseigner les Dettes Globales (Toutes les Dettes de L'EPE)"
    top.font = Font(name='Calibri', bold=True, size=15, color='000000ff')
    
    top = worksheet.cell (row=41, column=2)
    top.font = Font(color='000000ff')
    
    head1 = worksheet.cell (row=62, column=1)
    head1.value = ""
    
    for i in range(2, 8):
        cell = worksheet.cell(row=39, column=i)
        cell.fill = PatternFill(start_color="00cc99ff", end_color=("00cc99ff"), fill_type='solid')
        cell = worksheet.cell(row=48, column=i)
        cell.fill = PatternFill(start_color="00cc99ff", end_color=("00cc99ff"), fill_type='solid')
        cell = worksheet.cell(row=49, column=i)
        cell.fill = PatternFill(start_color="00cc99ff", end_color=("00cc99ff"), fill_type='solid')
        cell = worksheet.cell(row=50, column=i)
        cell.fill = PatternFill(start_color="00cc99ff", end_color=("00cc99ff"), fill_type='solid')
     
    for i in range(48, 51):        
        cell = worksheet.cell (row=i, column=5)
        cell.fill = PatternFill(start_color="00ffffff", end_color=("00ffffff"), fill_type='solid')
        
    cell = worksheet.cell (row=48, column=4)
    cell.fill = PatternFill(start_color="00ffffff", end_color=("00ffffff"), fill_type='solid')
    
    for i in range(2,8):
        line1 = worksheet.cell (row=41, column=i)
        line1.border =  Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), top=Side(border_style='thick'))
        line1 = worksheet.cell (row=42, column=i)
        line1.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line1 = worksheet.cell (row=47, column=i)
        line1.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line1 = worksheet.cell (row=48, column=i)
        line1.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line1 = worksheet.cell (row=49, column=i)
        line1.border =  Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), top=Side(border_style='thick'))
        line1 = worksheet.cell (row=50, column=i)
        line1.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    
    
    
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=6)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
        
    for i in range(37,51):   
        cell = worksheet.cell (row=i, column=7)
        cell.number_format = '0.00'
        
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=8)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=9)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    worksheet2 = workbook.create_sheet('Trésorerie', 1)
    
    
    
    tresorerie_queryset = Tresorerie.objects.filter(canva_id=pk)
    
    for c in Canva.objects.filter(id=pk):
        columns = ['SFCT', 'Comptes SCF', 'intitulé', 'Mois m-1', 'Mois m', 'Observations']

    row_num = 4
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet2.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for a in tresorerie_queryset:
        row_num += 1
        n=n+1
        row = [51, a.SCF, a.banques, a.moism1, a.moism, a.observation]
    
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet2.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')    
            
            if cell.value is None:
                cell.value=0
                
    for i in range(5,16):
        obs = worksheet2.cell (row=i, column=6)
        if obs.value==0:
            obs.value=""
            
    
        
    worksheet2.merge_cells('A1:C1')
    merged2 = worksheet2['A1']
    merged2b = worksheet2['C1']
    merged2c = worksheet2['B1']
    merged2.value = 'EPE : ANABIB Spa'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    
    
    worksheet2.merge_cells('A2:C2')
    merged2 = worksheet2['A2']
    merged2b = worksheet2['C2']
    merged2c = worksheet2['B2']
    for c in Canva.objects.filter(id=pk):
        d = c.mois      
        a = c.année
        merged2.value = 'Période :  '+d+'  '+str(a)
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='medium'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet2.merge_cells('D1:F1')
    merged2 = worksheet2['D1']
    merged2b = worksheet2['F1']
    merged2c = worksheet2['E1']
    merged2.value = 'Groupe IMETAL Spa'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet2.merge_cells('D2:F2')
    merged2 = worksheet2['D2']
    merged2b = worksheet2['F2']
    merged2c = worksheet2['E2']
    merged2.value = 'Tbg 2 : Trésorerie'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='medium'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet2.merge_cells('A4:C4')
    merged2 = worksheet2['A4']
    merged2b = worksheet2['C4']
    merged2.value = "Trésorerie (En MDA)"
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=20)
    
    worksheet2.merge_cells('A5:A15')
    merged2 = worksheet2['A5']
    merged2b = worksheet2['A15']
    merged2.value = "51"
    merged2.alignment = Alignment(horizontal="left", vertical="top")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet2.merge_cells('B5:C5')
    merged2 = worksheet2['B5']
    merged2b = worksheet2['C5']
    merged2.value = "Banques, Etablissement Financiers et Assimilés"
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    column_letter = get_column_letter(3)
    column_dimensions = worksheet2.column_dimensions[column_letter]
    column_dimensions.width = 45
    
    for i in range(4, 7):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet2.column_dimensions[column_letter]
        column_dimensions.width = 15
        
    for i in range(2, 7):
        cell = worksheet2.cell(row=5, column=i)
        cell.fill = PatternFill(start_color="00008080", end_color=("00008080"), fill_type='solid')
        
    for i in range(3, 7):
        cell = worksheet2.cell(row=11, column=i)
        cell.fill = PatternFill(start_color="0033cccc", end_color=("0033cccc"), fill_type='solid')
    
    cell = worksheet2.cell(row=13, column=3)
    cell.font = Font(color="00ff0000")
    
    for i in range(6,16):
        ag = worksheet2.cell (row=i, column=3)
        ag.alignment = Alignment('left')
    
    
    for i in range(12,16):
        ind1 = worksheet2.cell (row=i, column=3)
        ind1.alignment = Alignment(indent=3)
        ind2 = worksheet2.cell (row=i, column=2)
        ind2.alignment = Alignment('right')
         
    for i in range(6,12):
        ag = worksheet2.cell (row=i, column=2)
        ag.font = Font(name='Calibri', bold=True)
        ag2 = worksheet2.cell (row=i, column=3)
        ag2.font = Font(name='Calibri', bold=True)
        
    for i in range(4,7):
        ag = worksheet2.cell (row=4, column=i)
        ag.font = Font(name='Calibri', bold=True, size=12 )
        ag.alignment = Alignment(horizontal="center", vertical="center")
        
    for i in range(5,16):
        cell = worksheet2.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(5,16):
        cell = worksheet2.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for c in range(4,6):
        cell = worksheet2.cell (row=5, column=c)
        cell.value=""
        
    
        
        
        
    
    
    
    
    
    
    
    
    worksheet3 = workbook.create_sheet('Production physique', 2)
    
    
    
    production_queryset = Production.objects.filter(canva_id=pk)
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['Production physique', 'Unité', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 6
    
    for col_num, column_title in enumerate(columns, 2):
        cell = worksheet3.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for a in production_queryset:
        row_num += 1
        n=n+1
        if a.mois2 is None:
            a.mois2=0
        if a.mois1 is None or a.mois1==0:
            a.mois1=0
            row = [a.produit, a.unité, a.mois1, a.mois2, a.mois2-a.mois1, 100, a.finmois1, a.finmois2, a.ecart2, a.evolution2]
    
        
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet3.cell(row=row_num, column=col_num+1)
                cell.value = cell_value
                cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
                cell.alignment = Alignment('center')    
            
                if cell.value is None:
                    cell.value=0
        
        else:
            row = [a.produit, a.unité, a.mois1, a.mois2, a.mois2-a.mois1, (a.mois2-a.mois1)/a.mois1*100, a.finmois1, a.finmois2, a.ecart2, a.evolution2]
    
        
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet3.cell(row=row_num, column=col_num+1)
                cell.value = cell_value
                cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
                cell.alignment = Alignment('center')    
            
                if cell.value is None:
                    cell.value=0
            
            
    
    worksheet3.merge_cells('B1:D1')
    merged2 = worksheet3['B1']
    merged2b = worksheet3['D1']
    merged2c = worksheet3['C1']
    merged2.value = 'EPE : ANABIB Spa'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    
    
    worksheet3.merge_cells('B2:D2')
    merged2 = worksheet3['B2']
    merged2b = worksheet3['D2']
    merged2c = worksheet3['c2']
    for c in Canva.objects.filter(id=pk):
        d = c.mois      
        a = c.année
    merged2.value = 'Période :  '+d+'  '+str(a)
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='medium'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet3.merge_cells('E1:K1')
    merged2 = worksheet3['E1']
    merged2b = worksheet3['K1']
    merged2c = worksheet3['F1']
    merged2d = worksheet3['G1']
    merged2e = worksheet3['H1']
    merged2f = worksheet3['I1']
    merged2g = worksheet3['J1']
    merged2.value = 'Groupe IMETAL Spa'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2d.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2e.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2f.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2g.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet3.merge_cells('E2:K2')
    merged2 = worksheet3['E2']
    merged2b = worksheet3['K2']
    merged2c = worksheet3['F2']
    merged2d = worksheet3['G2']
    merged2e = worksheet3['H2']
    merged2f = worksheet3['I2']
    merged2g = worksheet3['J2']
    merged2.value = 'Tbg 2 : Production physique'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='medium'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2d.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2e.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2f.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2g.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet3.merge_cells('B5:B6')
    merged2 = worksheet3['B5']
    merged2b = worksheet3['B6']
    merged2.value = "Production physique"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=15)
    
    worksheet3.merge_cells('C5:C6')
    merged2 = worksheet3['C5']
    merged2b = worksheet3['C6']
    merged2.value = "Unité"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=15)
    
    
    worksheet3.merge_cells('D5:E5')
    merged2 = worksheet3['D5']
    merged2b = worksheet3['E5']
    merged2.value = "Mois"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=12)
    
    worksheet3.merge_cells('F5:G5')
    merged2 = worksheet3['F5']
    merged2b = worksheet3['G5']
    merged2.value = "Evolution"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    
    worksheet3.merge_cells('H5:I5')
    merged2 = worksheet3['H5']
    merged2b = worksheet3['I5']
    merged2.value = "Cumul"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=12)
    
    worksheet3.merge_cells('J5:K5')
    merged2 = worksheet3['J5']
    merged2b = worksheet3['K5']
    merged2.value = "Evolution"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    
    cell = worksheet3.cell (row=6, column=6)
    cell.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    cell = worksheet3.cell (row=6, column=7)
    cell.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    cell = worksheet3.cell (row=6, column=10)
    cell.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    cell = worksheet3.cell (row=6, column=11)
    cell.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    
    
    
    
    
    
    
    
    
    
    
    column_letter = get_column_letter(2)
    column_dimensions = worksheet3.column_dimensions[column_letter]
    column_dimensions.width = 35
    
    column_letter = get_column_letter(3)
    column_dimensions = worksheet3.column_dimensions[column_letter]
    column_dimensions.width = 15
    
    for i in range(4,12):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet3.column_dimensions[column_letter]
        column_dimensions.width = 20
    
    for i in range(8,30):
        ag = worksheet3.cell (row=i, column=2)
        ag.alignment = Alignment('left')
        ag.font = Font(name='Calibri', bold=True)
    
    
    for i in range(10,13):
        ind1 = worksheet3.cell (row=i, column=2)
        ind1.alignment = Alignment(indent=3)
        ind1.font = Font(name='Calibri', bold=False)
    
    for i in range(7,30):
        ag = worksheet3.cell (row=i, column=3)
        ag.font = Font(name='Calibri', bold=True)
        
    for i in range(2,12):
        ag = worksheet3.cell (row=7, column=i)
        ag.font = Font(name='Calibri', bold=True)
        ag.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        
    for c in range(4,12):
        cell = worksheet3.cell (row=7, column=c)
        cell.value=""
            
    
    
    
    
    
    
    
    
    
    
    
    
    
    worksheet4 = workbook.create_sheet('année', 3)
    for c in Canva.objects.filter(id=pk):
        allbilan_queryset = Bilan.objects.filter(canva__created__year=c.année, canva__created__lte=c.created, canva__site=c.site)
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['Agregat', 'SCF', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet4.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for b in allbilan_queryset:
        row_num += 1
        n=n+1
        row = [b.agregat, b.SCF, b.mois1, b.mois2]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet4.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')  
                 
            
            if cell.value is None:
                cell.value=0
        
    
    for i in range(1,5):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet4.column_dimensions[column_letter]
        column_dimensions.width = 35 
        
    for c in range(1,5):
        for l in range (2,314):
            cell = worksheet4.cell(row=l, column=c)
            if cell.value=="Vente de Marchandises":
                for d in range(1,5):
                    ligne = worksheet4.cell(row=l, column=d)
                    ligne.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
                 
    
    
    
    
    
            
    
    n=2
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=8, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    
    n=2
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=8, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26     

    n=3
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=9, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=3
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            cell1.value=0
        else:
            break
        cell = worksheet.cell(row=9, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26     

    n=4
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=10, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=4
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=10, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26    
    
    n=5
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=11, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=5
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            cell1.value=0
        else:
            break
        cell = worksheet.cell(row=11, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26  

    n=6
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=12, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=6
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=12, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26    
        
    n=7
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=13, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=7
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=13, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26    
        
    n=8
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=14, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=8
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=14, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26  

    n=9
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=15, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=9
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=15, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26  
        
    n=10
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=16, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=10
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=16, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26  
        
    n=11
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=17, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=11
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=17, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26  
        
    n=12
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=18, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=12
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=18, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=13
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=19, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=13
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=19, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=14
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=20, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=14
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=20, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=15
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=21, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=15
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=21, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=16
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=22, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=16
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=22, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26

    n=17
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=23, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=17
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=23, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=18
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=24, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=18
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=24, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=19
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=25, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=19
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=25, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=20
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=26, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=20
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=26, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=21
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=27, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=21
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=27, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=22
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=28, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=22
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=28, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26

    n=23
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=29, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=23
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=29, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=24
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=30, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=24
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=30, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=25
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=31, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=25
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=31, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=26
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=32, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=26
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=32, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        

    
    
    
    
    
    
    
    
    
    
    
    
    
    worksheet6 = workbook.create_sheet('année prod', 4)
    
    
    for c in Canva.objects.filter(id=pk):
        allprod_queryset = Production.objects.filter(canva__created__year=c.année, canva__created__lte=c.created, canva__site=c.site)
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['Agregat', 'SCF', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet6.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for b in allprod_queryset:
        row_num += 1
        n=n+1
        row = [b.produit, b.unité, b.mois1, b.mois2]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet6.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')  
                 
            
            if cell.value is None:
                cell.value=0

    for i in range(1,5):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet6.column_dimensions[column_letter]
        column_dimensions.width = 35 
        
    for c in range(1,5):
        for l in range (2,314):
            cell = worksheet6.cell(row=l, column=c)
            if cell.value=="PRODUIT":
                for d in range(1,5):
                    ligne = worksheet6.cell(row=l, column=d)
                    ligne.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')

                    
    n=3
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=8, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=3
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=8, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23    

    n=4
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=9, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=4
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=9, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23   

    n=5
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=10, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=5
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=10, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23   
    
    n=6
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=11, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=6
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=11, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23  

    n=7
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=12, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=7
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=12, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23    
        
    n=8
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=13, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=8
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=13, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23    
        
    n=9
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=14, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=9
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=14, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23 

    n=10
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=15, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=10
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=15, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23 
        
    n=11
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=16, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=11
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=16, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23  
        
    n=12
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=17, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=12
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=17, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23 
        
    n=13
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=18, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=13
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=18, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=14
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=19, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=14
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=19, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=15
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=20, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=15
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=20, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=16
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=21, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=16
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=21, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=17
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=22, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=17
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=22, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23

    n=18
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=23, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=18
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=23, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=19
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=24, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=19
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=24, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=20
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=25, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=20
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=25, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=21
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=26, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=21
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=26, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=22
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=27, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=22
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=27, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=23
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=28, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=23
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=28, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23

    n=24
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=29, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=24
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=29, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23


    for i in range(8,34):
        cell = worksheet.cell (row=i, column=8)
        if cell.value!=0:
            cell.number_format = '### ### ###'
            
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=9)
        if cell.value!=0:
            cell.number_format = '### ### ###'

     
    for l in range(8,34):
        cell1 = worksheet.cell (row=l, column=8)
        cell2 = worksheet.cell (row=l, column=9)
        cell = worksheet.cell (row=l, column=10)
        cell.value=cell2.value-cell1.value
        
    for l in range(8,34):
        cell1 = worksheet.cell (row=l, column=8)
        cell2 = worksheet.cell (row=l, column=9)
        cell = worksheet.cell (row=l, column=11)
        if cell1.value==0:
            cell.value=100
        else:
            cell.value=(cell2.value-cell1.value)/cell1.value*100
            
    for i in range(8,34):   
        cell = worksheet.cell (row=i, column=11)
        cell.number_format = '0.00'
        
    
    
    
    for i in range(8,34):
        cell = worksheet3.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(8,34):
        cell = worksheet3.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(8,30):
        cell = worksheet3.cell (row=i, column=6)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(8,30):
        cell = worksheet3.cell (row=i, column=8)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(8,30):
        cell = worksheet3.cell (row=i, column=9)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    
    for l in range(8,30):
        cell1 = worksheet3.cell (row=l, column=8)
        cell2 = worksheet3.cell (row=l, column=9)
        cell = worksheet3.cell (row=l, column=10)
        cell.value=cell2.value-cell1.value
        
    for l in range(8,30):
        cell1 = worksheet3.cell (row=l, column=8)
        cell2 = worksheet3.cell (row=l, column=9)
        cell = worksheet3.cell (row=l, column=11)
        if cell1.value==0:
            cell.value=100
        else:
            cell.value=(cell2.value-cell1.value)/cell1.value*100
            
    for i in range(8,30):   
        cell = worksheet3.cell (row=i, column=11)
        cell.number_format = '0.00'
        
    for i in range(8,30):   
        cell = worksheet3.cell (row=i, column=7)
        cell.number_format = '0.00'
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    for i in range(8,34):
        cell = worksheet3.cell (row=i, column=10)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
        
    
    
    
    workbook.save(response) 
    return response


        
    
    
   
    
    
    


class ExcelPageView(TemplateView):
    template_name = "excel_home.html"




class HomeView(TemplateView):
    template_name = 'home.html'
    

    
    


class CanvaCreateView(CreateView):
    model =Canva
    template_name = 'canva_create.html'
    fields = ['name', 'site', 'mois', 'année', 'created']
    def form_valid(self, form):
        messages.add_message(
            self.request,
            messages.SUCCESS,
            'Le Canvas a été ajouter'
        )
        
        return super().form_valid(form)
    
class CanvaDetailView(DetailView):
    model = Canva
    template_name = 'canva_detail.html'
    
class CanvaConsultlView(DetailView):
    model = Canva
    template_name = 'canva_consulter.html'
    

class CheckboxsEditView(SingleObjectMixin, FormView):
    model = Canva
    template_name = 'canva_valid.html'
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=Canva.objects.all())
        return super().get(request, *args, **kwargs)
    
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=Canva.objects.all())
        return super().post(request, *args, **kwargs)
    
    
    def get_form(self, form_class=None):
        return CheckboxsFormset(**self.get_form_kwargs(), instance=self.object)
    

    
    def form_valid(self, form):
        form.save()
        
        messages.add_message(
            self.request,
            messages.SUCCESS,
            'Canvas Enregistré'
            )
        
        return HttpResponseRedirect(self.get_success_url())
    
    
    def get_success_url(self):
        return reverse('bilans:canva_detail', kwargs={'pk': self.object.pk})

class BilansEditView(SingleObjectMixin, FormView):
    model = Canva
    template_name = 'bilan_edit.html'
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=Canva.objects.all())
        return super().get(request, *args, **kwargs)
    
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=Canva.objects.all())
        return super().post(request, *args, **kwargs)
    
    
    def get_form(self, form_class=None):
        return BilansFormset(**self.get_form_kwargs(), instance=self.object)
    

    
    def form_valid(self, form):
        form.save()
        
        messages.add_message(
            self.request,
            messages.SUCCESS,
            'Canvas Enregistré'
            )
        
        return HttpResponseRedirect(self.get_success_url())
    
    
    def get_success_url(self):
        return reverse('bilans:canva_detail', kwargs={'pk': self.object.pk})
    
    


class AutresEditView(SingleObjectMixin, FormView):
    model = Canva
    template_name = 'autre_edit.html'
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=Canva.objects.all())
        return super().get(request, *args, **kwargs)
    
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=Canva.objects.all())
        return super().post(request, *args, **kwargs)
    
    
    
    def get_form(self, form_class=None):
        return AutresFormset(**self.get_form_kwargs(), instance=self.object)
    
    
    
    
    def form_valid(self, form):
        form.save()
        
        messages.add_message(
            self.request,
            messages.SUCCESS,
            'Canvas Enregistré'
            )
        
        return HttpResponseRedirect(self.get_success_url())
    
    
    def get_success_url(self):
        return reverse('bilans:canva_detail', kwargs={'pk': self.object.pk})
        

        
        

class TresoreriesEditView(SingleObjectMixin, FormView):
    model = Canva
    template_name = 'tresorerie_edit.html'
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=Canva.objects.all())
        return super().get(request, *args, **kwargs)
    
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=Canva.objects.all())
        return super().post(request, *args, **kwargs)
    
    
    
    def get_form(self, form_class=None):
        return TresoreriesFormset(**self.get_form_kwargs(), instance=self.object)
    
    
    
    
    def form_valid(self, form):
        form.save()
        
        messages.add_message(
            self.request,
            messages.SUCCESS,
            'Canvas Enregistré'
            )
        
        return HttpResponseRedirect(self.get_success_url())
    
    
    def get_success_url(self):
        return reverse('bilans:canva_detail', kwargs={'pk': self.object.pk})
    

class ProductionsEditView(SingleObjectMixin, FormView):
    model = Canva
    template_name = 'production_edit.html'
    
    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=Canva.objects.all())
        return super().get(request, *args, **kwargs)
    
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=Canva.objects.all())
        return super().post(request, *args, **kwargs)
    
    
    
    def get_form(self, form_class=None):
        return ProductionsFormset(**self.get_form_kwargs(), instance=self.object)
    
    
    
    
    def form_valid(self, form):
        form.save()
        
        messages.add_message(
            self.request,
            messages.SUCCESS,
            'Canvas Enregistré'
            )
        
        return HttpResponseRedirect(self.get_success_url())
    
    
    def get_success_url(self):
        return reverse('bilans:canva_detail', kwargs={'pk': self.object.pk})


@login_required
@allowed_users(allowed_roles=['CDG', 'SG', 'admin', 'DG', 'ALTUMET', 'IRRAGRIS', 'PAF', 'PTS', 'PTTP'])
def index(request):
        now = timezone.now    
        DG_list = Canva.objects.filter(site="DG").order_by('created')
        ALTUMET_list = Canva.objects.filter(site="ALTUMET").order_by('created')
        IRRAGRIS_list = Canva.objects.filter(site="IRRAGRIS").order_by('created')
        PAF_list = Canva.objects.filter(site="PAF").order_by('created')
        PTS_list = Canva.objects.filter(site="PTS").order_by('created')
        PTTP_list = Canva.objects.filter(site="PTTP").order_by('created')
        
       
        
            
        context_dict = {'DG': DG_list,'ALTUMET':ALTUMET_list, 'IRRAGRIS':IRRAGRIS_list,'PAF': PAF_list, 'PTS': PTS_list, 'PTTP': PTTP_list, 'now': now}
        return render(request, 'canva_list.html', context=context_dict)

@login_required
@allowed_users(allowed_roles=['CDG', 'SG', 'admin', 'DG', 'ALTUMET', 'IRRAGRIS', 'PAF', 'PTS', 'PTTP'])
def index2(request):
        now = timezone.now    
        
        DG_list = Canva.objects.filter(site="DG").order_by('created')
        
        context_dict = {'DG': DG_list, 'now': now}
        return render(request, 'canva_list2.html', context=context_dict)
        
        
        
        
        
        
        
































        
        
        
        
def export_somme_xls(request, pk):    
   
    bilan_queryset = Bilan.objects.filter(canva_id=pk)
    
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    for c in Canva.objects.filter(id=pk):
        d = c.mois      
        a = c.année

    response['Content-Disposition'] = 'attachment; filename=Sommes Agregats Mensuels - {mois} - {année}.xlsx'.format(mois=d, année=a)
    
    
    workbook = Workbook()
    
    worksheet = workbook.active
    worksheet.title = 'TB SCF Ecofie'
    
    
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['N°', 'Agregat', 'SCF', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 7
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for b in bilan_queryset:
        row_num += 1
        n=n+1
        
        
        row = [n,
            b.agregat, 
            b.SCF, 
            0,
            0,
            0, 
            0, 
            0, 
            0, 
            0, 
            0,]
    
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')  
                 
            
            if cell.value is None:
                cell.value=0
        
        
                    
    
    
    
    
    top = worksheet.cell (row=1, column=2)
    top.value = "TABLEAU DE BORD MENSUEL"
    top.font = Font(name='Calibri', bold=True, size=20, color='000000ff')

    b = 6
    for i in range(14,34):
        ag = worksheet.cell (row=i, column=1)
        ag.value = b
        b = b+1
    
    
    
    
    
    

    
    
    
    
    top2 = worksheet.cell (row=1, column=8)
    for c in Canva.objects.filter(id=pk):
        top2.value = c.mois
        top2.font = Font(name='Calibri', bold=True, size=20)
    
    top3 = worksheet.cell (row=1, column=10)
    for c in Canva.objects.filter(id=pk):
        top3.value = c.année
        top3.font = Font(name='Calibri', bold=True, size=20)
    
    top4 = worksheet.cell (row=3, column=2)
    top4.value = "Groupe"
    top4.border = Border(bottom=Side(border_style='thick'), left=Side(border_style='thick'), top=Side(border_style='thick'), right=Side(border_style='thick'))
    top4.font = Font(name='Calibri', bold=True, size=13)
    top4.alignment = Alignment(horizontal='center')
    
    top5 = worksheet.cell (row=4, column=2)
    top5.value = "EPE"
    top5.alignment = Alignment(horizontal='center')
    top5.border = Border(left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    top5.font = Font(name='Calibri', bold=True, size=13)
    
    top6 = worksheet.cell (row=3, column=5)
    top6.value = "IMETAL"
    top6.border = Border(bottom=Side(border_style='thick'), top=Side(border_style='thick'))
    top6.font = Font(name='Calibri', bold=True, size=13, color='000000ff')
    
    top7 = worksheet.cell (row=4, column=5)
    top7.value = "ANABIB"
    top7.border = Border(bottom=Side(border_style='thick'))
    top7.font = Font(name='Calibri', bold=True, size=13, color='000000ff')
    
    top8 = worksheet.cell (row=5, column=10)
    top8.value = "En millions de DA"
    top8.font = Font(name='Calibri', bold=True, size=13, color='00ff0000')
    
    vide1 = worksheet.cell (row=3, column=3)
    vide1.border = Border(bottom=Side(border_style='thick'), top=Side(border_style='thick'))
    vide2 = worksheet.cell (row=3, column=4)
    vide2.border = Border(bottom=Side(border_style='thick'), top=Side(border_style='thick'))
    vide3 = worksheet.cell (row=3, column=6)
    vide3.border = Border(bottom=Side(border_style='thick'), top=Side(border_style='thick'), right=Side(border_style='thick'))
    vide4 = worksheet.cell (row=4, column=3)
    vide4.border = Border(bottom=Side(border_style='thick'))
    vide5 = worksheet.cell (row=4, column=4)
    vide5.border = Border(bottom=Side(border_style='thick'))
    vide6 = worksheet.cell (row=4, column=6)
    vide6.border = Border(bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    
    
    
    
    
    
    
    
    column_letter = get_column_letter(1)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 5
    
    
    column_letter = get_column_letter(2)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 45
    
    column_letter = get_column_letter(3)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 13
    
    for i in range(4,12):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 17
    
    column_letter = get_column_letter(8)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 20
    
    column_letter = get_column_letter(9)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 20
    
    
    
    for i in range(8,34):
        ag = worksheet.cell (row=i, column=2)
        ag.alignment = Alignment('left')
    
    
    
    ind1 = worksheet.cell (row=20, column=2)
    ind1.alignment = Alignment(indent=2)
    
    ind2 = worksheet.cell (row=21, column=2)
    ind2.alignment = Alignment(indent=5)
    
    ind3 = worksheet.cell (row=22, column=2)
    ind3.alignment = Alignment(indent=5)
    
    worksheet.merge_cells('A13:A14')
    merged1 = worksheet['A13']
    merged1.value = "6"
    merged1.alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet.merge_cells('A6:A7')
    merged2 = worksheet['A6']
    merged2b = worksheet['A7']
    merged2.value = 'N°'
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('B6:B7')
    merged3 = worksheet['B6']
    merged3b = worksheet['B7']
    merged3.value = 'AGREGATS'
    merged3.alignment = Alignment(horizontal="center", vertical="center")
    merged3.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged3b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged3.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged3.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('C6:C7')
    merged4 = worksheet['C6']
    merged4b = worksheet['C7']
    merged4.value = 'Comptes SCF'
    merged4.alignment = Alignment(horizontal="center", vertical="center")
    merged4.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged4b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged4.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged4.font = Font(name='Calibri', bold=True, size=11)
    
    worksheet.merge_cells('D6:E6')
    merged5 = worksheet['D6']
    merged5b = worksheet['E6']
    merged5.value = 'Mois'
    merged5.alignment = Alignment(horizontal="center", vertical="center")
    merged5.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged5b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged5.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged5.font = Font(name='Calibri', bold=True, size=13)
    
    
    worksheet.merge_cells('F6:G6')
    merged6 = worksheet['F6']
    merged6b = worksheet['G6']
    merged6.value = 'Evolution'
    merged6.alignment = Alignment(horizontal="center", vertical="center")
    merged6.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged6b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged6.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged6.font = Font(name='Calibri', bold=True, size=13, color='000000ff')
    
    worksheet.merge_cells('H6:I6')
    merged7 = worksheet['H6']
    merged7b = worksheet['I6']
    merged7.value = 'Cumul'
    merged7.alignment = Alignment(horizontal="center", vertical="center")
    merged7.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged7b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged7.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged7.font = Font(name='Calibri', bold=True, size=13)
    
    
    
    worksheet.merge_cells('J6:K6')
    merged8 = worksheet['J6']
    merged8b = worksheet['K6']
    merged8.value = 'Evolution'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=13, color='000000ff')
    
    head1 = worksheet.cell (row=7, column=6)
    head1.font = Font(color='000000ff')
    
    head2 = worksheet.cell (row=7, column=7)
    head2.font = Font(color='000000ff')
    
    head3 = worksheet.cell (row=7, column=10)
    head3.font = Font(color='000000ff')
    
    head4 = worksheet.cell (row=7, column=11)
    head4.font = Font(color='000000ff')

    
    for i in range(1,12):
        line1 = worksheet.cell (row=13, column=i)
        line1.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'))
        line1.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        line2 = worksheet.cell (row=14, column=i)
        line2.border =  Border(bottom=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), top=Side(border_style='medium'))
        line2.fill = PatternFill(start_color="00ffcc99", end_color=("00ffcc99"), fill_type='solid')
        line3 = worksheet.cell (row=18, column=i)
        line3.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line3.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        line4 = worksheet.cell (row=25, column=i)
        line4.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line4.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        line5 = worksheet.cell (row=26, column=i)
        line5.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line5.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        line6 = worksheet.cell (row=29, column=i)
        line6.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line6.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        line7 = worksheet.cell (row=33, column=i)
        line7.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line7.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    
    
    
    
    test1 = worksheet.cell (row=8, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=4)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=4)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=4)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=4)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=5)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=5)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=5)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=5)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=5)
    sum1.value = int(value1+value2+value3+value4+value5) 
    
    test1 = worksheet.cell (row=8, column=6)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=6)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=6)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=6)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=6)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    
    
    
    
    test1 = worksheet.cell (row=13, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=13, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=13, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    
    
    
    
    
    test1 = worksheet.cell (row=8, column=8)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=8)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=8)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=8)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=8)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=9)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=9)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=9)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=9)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=9)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=10)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=10)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=10)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=10)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=10)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=11)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=11)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=11)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=11)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=11)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=11)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    
    
    test1 = worksheet.cell (row=13, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=4)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=4)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=4)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=13, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=5)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=5)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=5)
    sum1.value = int(value1+value2+value3+value4)
    sum1.number_format = '### ### ###'
    
    test1 = worksheet.cell (row=13, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=6)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=6)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=6)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=18, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=18, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=18, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    test1 = worksheet.cell (row=13, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=8)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=8)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=8)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=13, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=9)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=9)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=9)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=13, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=10)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=10)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=10)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=19, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=4)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=4)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=5)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=5)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=6)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=6)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=25, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=25, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=25, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    test1 = worksheet.cell (row=19, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=8)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=8)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=9)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=9)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=10)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=10)
    sum1.value = int(value1+value2+value3)
    
    
    cell = worksheet.cell (row=8, column=5)
    cell.number_format = '### ### ###'
    
    
    
    test1 = worksheet.cell (row=18, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=4)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=4)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=5)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=6)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=6)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=8)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=8)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=9)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=9)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=10)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=10)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=26, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=26, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    test1 = worksheet.cell (row=26, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=4)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=4)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=5)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=5)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=6)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=6)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=8)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=8)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=9)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=9)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=10)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=10)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=29, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=29, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=29, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    
    
    
    
    
    
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=6)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
        
    for i in range(8,34):   
        cell = worksheet.cell (row=i, column=7)
        cell.number_format = '0.00'
    
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=8)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=9)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
           
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=10)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    


    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    worksheet4 = workbook.create_sheet('somme', 3)
    for c in Canva.objects.filter(id=pk):
        allbilan_queryset = Bilan.objects.filter(canva__mois=c.mois, canva__année=c.année)
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['Agregat', 'SCF', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet4.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    
    for b in allbilan_queryset:
        row_num += 1
        n=n+1
        row = [b.agregat, b.SCF, b.mois1, b.mois2]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet4.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')  
                 
            
            if cell.value is None:
                cell.value=0


    for i in range(1,160):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet4.column_dimensions[column_letter]
        column_dimensions.width = 20


    n=2
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=8, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    
    n=2
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=8, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26     

    n=3
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=9, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=3
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            cell1.value=0
        else:
            break
        cell = worksheet.cell(row=9, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26     

    n=4
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=10, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=4
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=10, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26    
    
    n=5
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=11, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=5
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            cell1.value=0
        else:
            break
        cell = worksheet.cell(row=11, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26  

    n=6
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=12, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=6
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=12, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26    
        
    n=7
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=13, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=7
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=13, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26    
        
    n=8
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=14, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=8
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=14, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26  

    n=9
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=15, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=9
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=15, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26  
        
    n=10
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=16, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=10
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=16, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26  
        
    n=11
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=17, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=11
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=17, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26  
        
    n=12
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=18, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=12
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=18, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=13
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=19, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=13
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=19, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=14
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=20, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=14
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=20, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=15
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=21, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=15
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=21, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=16
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=22, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=16
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=22, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26

    n=17
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=23, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=17
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=23, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=18
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=24, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=18
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=24, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=19
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=25, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=19
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=25, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=20
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=26, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=20
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=26, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=21
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=27, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=21
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=27, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=22
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=28, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=22
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=28, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26

    n=23
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=29, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=23
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=29, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=24
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=30, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=24
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=30, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=25
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=31, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=25
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=31, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=26
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=32, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=26
    for i in range(1,13):
        cell1 = worksheet4.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=32, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
        
    autre_queryset = Autre.objects.filter(canva_id=pk)
    
    for c in Canva.objects.filter(id=pk):
        columns = ['N°', 'Autres Agregats', 'Comptes SCF', 'a la fin Mois N-1', 'a la fin Mois N', 'Ecart en Valeur', '% evolution', 'Observations:']

    row_num = 36
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for a in autre_queryset:
        row_num += 1
        n=n+1
        
        
        row = [n, a.autreag, a.SCF, 0, 0, 0, 0,0]
    
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')    
            
            if cell.value is None:
                cell.value=0
        
       
        
        
    
    
    
            
    
    
    
    
    for i in range(37,51):
        ag = worksheet.cell (row=i, column=2)
        ag.alignment = Alignment('left')
    
    ind1 = worksheet.cell (row=38, column=2)
    ind1.alignment = Alignment(indent=2)
    ind1.font = Font(color='00008000')
    
    ind1 = worksheet.cell (row=38, column=3)
    ind1.font = Font(color='00008000')
    
    ind2 = worksheet.cell (row=39, column=2)
    ind2.alignment = Alignment(indent=5)
    
    ind3 = worksheet.cell (row=40, column=2)
    ind3.alignment = Alignment(indent=8)
    ind3.font = Font(color='00ff0000')
    
    ind3 = worksheet.cell (row=40, column=3)
    ind3.font = Font(color='00ff0000')
    
    ind4 = worksheet.cell (row=42, column=2)
    ind4.alignment = Alignment(indent=5)
    
    ind5 = worksheet.cell (row=44, column=2)
    ind5.alignment = Alignment(indent=2)
    
    ind6 = worksheet.cell (row=45, column=2)
    ind6.alignment = Alignment(indent=5)
    
    ind7 = worksheet.cell (row=46, column=2)
    ind7.alignment = Alignment(indent=5)
    
    ind8 = worksheet.cell (row=48, column=2)
    ind8.alignment = Alignment(indent=5)
    
    ind9 = worksheet.cell (row=50, column=2)
    ind9.alignment = Alignment(indent=5)
    
    worksheet.merge_cells('B35:B36')
    merged8 = worksheet['B35']
    merged8b = worksheet['B36']
    merged8.value = 'Autres Agrégats'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('C35:C36')
    merged8 = worksheet['C35']
    merged8b = worksheet['C36']
    merged8.value = 'Comptes SCF'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=11)
    
    worksheet.merge_cells('D35:E35')
    merged8 = worksheet['D35']
    merged8b = worksheet['E35']
    merged8.value = 'Cumul'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('F35:G35')
    merged8 = worksheet['F35']
    merged8b = worksheet['G35']
    merged8.value = 'Evolution'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=13, color='000000ff')
    
    worksheet.merge_cells('H35:K36')
    merged8 = worksheet['H35']
    merged8b = worksheet['K36']
    merged8.value = 'Observations :'
    merged8.alignment = Alignment(horizontal="left", vertical="top")
    merged8.border =  Border(left=Side(border_style='thick'))
    merged8.fill = PatternFill(start_color="0099cc00", end_color=("0099cc00"), fill_type='solid')
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A35:A36')
    merged8 = worksheet['A35']
    merged8b = worksheet['A36']
    merged8.value = ''
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A37:A40')
    merged8 = worksheet['A37']
    merged8b = worksheet['A40']
    merged8.value = '25'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A41:A42')
    merged8 = worksheet['A41']
    merged8b = worksheet['A42']
    merged8.value = '26'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A43:A46')
    merged8 = worksheet['A43']
    merged8b = worksheet['A46']
    merged8.value = '27'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A47:A48')
    merged8 = worksheet['A47']
    merged8b = worksheet['A48']
    merged8.value = '28'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet.merge_cells('A49:A50')
    merged8 = worksheet['A49']
    merged8b = worksheet['A50']
    merged8.value = '29'
    merged8.alignment = Alignment(horizontal="center", vertical="center")
    merged8.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8b.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged8.font = Font(name='Calibri', bold=True, size=13)
    
    
    
    
    for i in range(8, 12):
        for n in range(37, 51):
            ob = worksheet.cell (row=n, column=i)
            ob.alignment = Alignment('left')
            ob.alignment = Alignment(horizontal="left", vertical="top")
            ob.fill = PatternFill(start_color="0099cc00", end_color=("0099cc00"), fill_type='solid')
            ob.font = Font(name='Calibri', bold=True, size=11)
            ob.border =  Border(left=Side(border_style='none'))
            if ob.value==0:
                ob.value=" "
        
    
    head1 = worksheet.cell (row=36, column=6)
    head1.font = Font(color='000000ff')
    
    head1 = worksheet.cell (row=36, column=7)
    head1.font = Font(color='000000ff')
    
    top = worksheet.cell (row=52, column=1)
    top.value = "* Veuillez renseigner les Dettes Globales (Toutes les Dettes de L'EPE)"
    top.font = Font(name='Calibri', bold=True, size=15, color='000000ff')
    
    top = worksheet.cell (row=41, column=2)
    top.font = Font(color='000000ff')
    
    head1 = worksheet.cell (row=62, column=1)
    head1.value = ""
    
    for i in range(2, 8):
        cell = worksheet.cell(row=39, column=i)
        cell.fill = PatternFill(start_color="00cc99ff", end_color=("00cc99ff"), fill_type='solid')
        cell = worksheet.cell(row=48, column=i)
        cell.fill = PatternFill(start_color="00cc99ff", end_color=("00cc99ff"), fill_type='solid')
        cell = worksheet.cell(row=49, column=i)
        cell.fill = PatternFill(start_color="00cc99ff", end_color=("00cc99ff"), fill_type='solid')
        cell = worksheet.cell(row=50, column=i)
        cell.fill = PatternFill(start_color="00cc99ff", end_color=("00cc99ff"), fill_type='solid')
     
    for i in range(48, 51):        
        cell = worksheet.cell (row=i, column=5)
        cell.fill = PatternFill(start_color="00ffffff", end_color=("00ffffff"), fill_type='solid')
        
    cell = worksheet.cell (row=48, column=4)
    cell.fill = PatternFill(start_color="00ffffff", end_color=("00ffffff"), fill_type='solid')
    
    for i in range(2,8):
        line1 = worksheet.cell (row=41, column=i)
        line1.border =  Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), top=Side(border_style='thick'))
        line1 = worksheet.cell (row=42, column=i)
        line1.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line1 = worksheet.cell (row=47, column=i)
        line1.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line1 = worksheet.cell (row=48, column=i)
        line1.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        line1 = worksheet.cell (row=49, column=i)
        line1.border =  Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), top=Side(border_style='thick'))
        line1 = worksheet.cell (row=50, column=i)
        line1.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    
    
    
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=6)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
        
    for i in range(37,51):   
        cell = worksheet.cell (row=i, column=7)
        cell.number_format = '0.00'
        
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=8)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=9)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'



    
    
    
    
    
    
    
    
    
    
    
    
    
    worksheet5 = workbook.create_sheet('cumulsomme', 4)
    for c in Canva.objects.filter(id=pk):
        bilanCS_queryset = Bilan.objects.filter(canva__created__year=c.année, canva__created__lte=c.created)
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['Agregat', 'SCF', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet5.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for b in bilanCS_queryset:
        row_num += 1
        n=n+1
        row = [b.agregat, b.SCF, b.mois1, b.mois2]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet5.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')  
                 
            
            if cell.value is None:
                cell.value=0
        
    
    for i in range(1,5):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet5.column_dimensions[column_letter]
        column_dimensions.width = 35 
        
    for c in range(1,5):
        for l in range (2,1880):
            cell = worksheet5.cell(row=l, column=c)
            if cell.value=="Vente de Marchandises":
                for d in range(1,5):
                    ligne = worksheet5.cell(row=l, column=d)
                    ligne.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
                    
    n=2
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=8, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    
    n=2
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=8, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26     

    n=3
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=9, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=3
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            cell1.value=0
        else:
            break
        cell = worksheet.cell(row=9, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26     

    n=4
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=10, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=4
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=10, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26    
    
    n=5
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=11, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=5
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            cell1.value=0
        else:
            break
        cell = worksheet.cell(row=11, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26  

    n=6
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=12, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=6
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=12, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26    
        
    n=7
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=13, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=7
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=13, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26    
        
    n=8
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=14, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=8
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=14, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26  

    n=9
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=15, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=9
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=15, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26  
        
    n=10
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=16, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=10
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=16, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26  
        
    n=11
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=17, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=11
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=17, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26  
        
    n=12
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=18, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=12
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=18, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=13
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=19, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=13
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=19, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=14
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=20, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=14
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=20, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=15
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=21, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=15
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=21, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=16
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=22, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=16
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=22, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26

    n=17
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=23, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=17
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=23, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=18
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=24, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=18
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=24, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=19
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=25, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=19
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=25, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=20
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=26, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=20
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=26, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=21
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=27, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=21
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=27, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=22
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=28, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=22
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=28, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26

    n=23
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=29, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=23
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=29, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=24
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=30, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=24
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=30, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=25
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=31, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=25
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=31, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    n=26
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=32, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+26
    
    n=26
    for i in range(1,75):
        cell1 = worksheet5.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=32, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+26
        
    test1 = worksheet.cell (row=8, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=4)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=4)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=4)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=4)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=5)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=5)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=5)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=5)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=5)
    sum1.value = int(value1+value2+value3+value4+value5) 
    
    test1 = worksheet.cell (row=8, column=6)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=6)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=6)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=6)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=6)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    
    
    
    
    test1 = worksheet.cell (row=13, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=13, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=13, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    test1 = worksheet.cell (row=8, column=8)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=8)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=8)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=8)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=8)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=9)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=9)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=9)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=9)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=9)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=10)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=10)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=10)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=10)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=10)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    test1 = worksheet.cell (row=8, column=11)
    value1 = test1.value
    test2 = worksheet.cell (row=9, column=11)
    value2 = test2.value
    test3 = worksheet.cell (row=10, column=11)
    value3 = test3.value
    test4 = worksheet.cell (row=11, column=11)
    value4 = test4.value
    test5 = worksheet.cell (row=12, column=11)
    value5 = test5.value
    sum1 = worksheet.cell (row=13, column=11)
    sum1.value = int(value1+value2+value3+value4+value5)
    
    
    
    test1 = worksheet.cell (row=13, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=4)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=4)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=4)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=13, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=5)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=5)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=5)
    sum1.value = int(value1+value2+value3+value4)
    sum1.number_format = '### ### ###'
    
    test1 = worksheet.cell (row=13, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=6)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=6)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=6)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=18, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=18, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=18, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    test1 = worksheet.cell (row=13, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=8)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=8)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=8)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=13, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=9)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=9)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=9)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=13, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=15, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=16, column=10)
    value3 = test3.value
    test4 = worksheet.cell (row=17, column=10)
    value4 = test4.value
    sum1 = worksheet.cell (row=18, column=10)
    sum1.value = int(value1+value2+value3+value4)
    
    test1 = worksheet.cell (row=19, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=4)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=4)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=5)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=5)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=6)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=6)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=25, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=25, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=25, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    test1 = worksheet.cell (row=19, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=8)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=8)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=9)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=9)
    sum1.value = int(value1+value2+value3)
    
    test1 = worksheet.cell (row=19, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=23, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=24, column=10)
    value3 = test3.value
    sum1 = worksheet.cell (row=25, column=10)
    sum1.value = int(value1+value2+value3)
    
    
    cell = worksheet.cell (row=8, column=5)
    cell.number_format = '### ### ###'
    
    
    
    test1 = worksheet.cell (row=18, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=4)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=4)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=5)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=6)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=6)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=8)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=8)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=9)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=9)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=18, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=25, column=10)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=10)
    sum1.value = int(value1-value2)
    
    test1 = worksheet.cell (row=26, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=26, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=26, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    
    
    test1 = worksheet.cell (row=26, column=4)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=4)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=4)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=4)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=5)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=5)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=5)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=5)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=6)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=6)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=6)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=6)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=8)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=8)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=8)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=8)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=9)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=9)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=9)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=9)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=26, column=10)
    value1 = int(test1.value)
    test2 = worksheet.cell (row=27, column=10)
    value2 = test2.value
    test3 = worksheet.cell (row=28, column=10)
    value3 = test3.value
    sum1 = worksheet.cell (row=29, column=10)
    sum1.value = int(value1-value2-value3)
    
    test1 = worksheet.cell (row=29, column=4)
    value1 = test1.value
    test2 = worksheet.cell (row=29, column=5)
    value2 = test2.value
    sum1 = worksheet.cell (row=29, column=7)
    if value1==0:
        sum1.value = 0
    else:
        sum1.value = (int(value2)-int(value1))/int(value1)*100
    

    for i in range(8,34):
        cell1 = worksheet.cell (row=i, column=4)
        cell2 = worksheet.cell (row=i, column=5)
        value1=cell1.value
        value2=cell2.value
        cell = worksheet.cell (row=i, column=6)
        cell.value=int(value2-value1)
        
    for i in range(8,34):
        cell1 = worksheet.cell (row=i, column=4)
        cell2 = worksheet.cell (row=i, column=5)
        value1=cell1.value
        value2=cell2.value
        cell = worksheet.cell (row=i, column=7)
        if value1==0:
            cell.value=100
        else:
            cell.value=(int(value2)-int(value1))/int(value1)*100
            
            
    for i in range(8,34):
        cell1 = worksheet.cell (row=i, column=8)
        cell2 = worksheet.cell (row=i, column=9)
        value1=cell1.value
        value2=cell2.value
        cell = worksheet.cell (row=i, column=10)
        cell.value=int(value2-value1)
        
    for i in range(8,34):
        cell1 = worksheet.cell (row=i, column=8)
        cell2 = worksheet.cell (row=i, column=9)
        value1=cell1.value
        value2=cell2.value
        cell = worksheet.cell (row=i, column=11)
        if value1==0:
            cell.value=100
        else:
            cell.value=(int(value2)-int(value1))/int(value1)*100        
        

    
    
   
        
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=6)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
        
    for i in range(8,34):   
        cell = worksheet.cell (row=i, column=7)
        cell.number_format = '0.00'
    
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=8)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=9)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
           
    for i in range(8,34):
        cell = worksheet.cell (row=i, column=10)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(8,34):   
        cell = worksheet.cell (row=i, column=11)
        cell.number_format = '0.00'
    
    
    
    
    worksheet6 = workbook.create_sheet('autre somme', 5)
    for c in Canva.objects.filter(id=pk):
        autreS_queryset = Autre.objects.filter(canva__mois=c.mois, canva__année=c.année)
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['Agregat', 'SCF', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet6.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    
    for b in autreS_queryset:
        row_num += 1
        n=n+1
        row = [b.autreag, b.SCF, b.finmoisn1, b.finmoisn]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet6.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')  
                 
            
            if cell.value is None:
                cell.value=0
    for i in range(1,5):           
        column_letter = get_column_letter(i)
        column_dimensions = worksheet6.column_dimensions[column_letter]
        column_dimensions.width = 30
    
    for c in range(1,5):
        for l in range (2,1880):
            cell = worksheet6.cell(row=l, column=c)
            if cell.value=="Disponibilités et assimilés":
                for d in range(1,5):
                    ligne = worksheet6.cell(row=l, column=d)
                    ligne.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    
    
    n=2
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=37, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
        
    
    n=2
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=37, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14     

    n=3
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=38, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=3
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            cell1.value=0
        else:
            break
        cell = worksheet.cell(row=38, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14    

    n=4
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=39, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=4
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=39, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14   
    
    n=5
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=40, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=5
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            cell1.value=0
        else:
            break
        cell = worksheet.cell(row=40, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14  

    n=6
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=41, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=6
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=41, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14
        
    n=7
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=42, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=7
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=42, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14    
        
    n=8
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=43, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=8
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=43, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14 

    n=9
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=44, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=9
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=44, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14  
        
    n=10
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=45, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=10
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=45, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14  
        
    n=11
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=46, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=11
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=46, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14  
        
    n=12
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=47, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=12
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=47, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=13
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=48, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=13
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=48, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14
        
    n=14
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=49, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=14
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=49, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14
        
    n=15
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=50, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+14
    
    n=15
    for i in range(1,13):
        cell1 = worksheet6.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet.cell(row=50, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+14
        
    
    
    for i in range(37,51):
        test1 = worksheet.cell (row=i, column=4)
        value1 = test1.value
        test2 = worksheet.cell (row=i, column=5)
        value2 = test2.value
        sum1 = worksheet.cell (row=i, column=6)
        sum1.value = int(value2-value1)
        eva1 = worksheet.cell (row=i, column=7)
        if value1==0:
            eva1.value = 100
        else:
            eva1.value = (int(value2)-int(value1))/int(value1)*100
        
        
        
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=6)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
        
    for i in range(37,51):   
        cell = worksheet.cell (row=i, column=7)
        cell.number_format = '0.00'
        
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=8)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(37,51):
        cell = worksheet.cell (row=i, column=9)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
        
        
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    worksheet2 = workbook.create_sheet('Trésorerie', 1)
    
    
    
    tresorerieS_queryset = Tresorerie.objects.filter(canva_id=pk)
    
    for c in Canva.objects.filter(id=pk):
        columns = ['SFCT', 'Comptes SCF', 'intitulé', 'Mois m-1', 'Mois m', 'Observations']

    row_num = 4
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet2.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for a in tresorerieS_queryset:
        row_num += 1
        n=n+1
        row = [51, a.SCF, a.banques, 0, 0, 0]
    
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet2.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')    
            
            if cell.value is None:
                cell.value=0
                
    for i in range(5,16):
        obs = worksheet2.cell (row=i, column=6)
        if obs.value==0:
            obs.value=""
            
    
        
    worksheet2.merge_cells('A1:C1')
    merged2 = worksheet2['A1']
    merged2b = worksheet2['C1']
    merged2c = worksheet2['B1']
    merged2.value = 'EPE : ANABIB Spa'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    
    
    worksheet2.merge_cells('A2:C2')
    merged2 = worksheet2['A2']
    merged2b = worksheet2['C2']
    merged2c = worksheet2['B2']
    for c in Canva.objects.filter(id=pk):
        d = c.mois      
        a = c.année
        merged2.value = 'Période :  '+d+'  '+str(a)
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='medium'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet2.merge_cells('D1:F1')
    merged2 = worksheet2['D1']
    merged2b = worksheet2['F1']
    merged2c = worksheet2['E1']
    merged2.value = 'Groupe IMETAL Spa'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet2.merge_cells('D2:F2')
    merged2 = worksheet2['D2']
    merged2b = worksheet2['F2']
    merged2c = worksheet2['E2']
    merged2.value = 'Tbg 2 : Trésorerie'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='medium'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet2.merge_cells('A4:C4')
    merged2 = worksheet2['A4']
    merged2b = worksheet2['C4']
    merged2.value = "Trésorerie (En MDA)"
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=20)
    
    worksheet2.merge_cells('A5:A15')
    merged2 = worksheet2['A5']
    merged2b = worksheet2['A15']
    merged2.value = "51"
    merged2.alignment = Alignment(horizontal="left", vertical="top")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet2.merge_cells('B5:C5')
    merged2 = worksheet2['B5']
    merged2b = worksheet2['C5']
    merged2.value = "Banques, Etablissement Financiers et Assimilés"
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    column_letter = get_column_letter(3)
    column_dimensions = worksheet2.column_dimensions[column_letter]
    column_dimensions.width = 45
    
    for i in range(4, 7):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet2.column_dimensions[column_letter]
        column_dimensions.width = 15
        
    for i in range(2, 7):
        cell = worksheet2.cell(row=5, column=i)
        cell.fill = PatternFill(start_color="00008080", end_color=("00008080"), fill_type='solid')
        
    for i in range(3, 7):
        cell = worksheet2.cell(row=11, column=i)
        cell.fill = PatternFill(start_color="0033cccc", end_color=("0033cccc"), fill_type='solid')
    
    cell = worksheet2.cell(row=13, column=3)
    cell.font = Font(color="00ff0000")
    
    for i in range(6,16):
        ag = worksheet2.cell (row=i, column=3)
        ag.alignment = Alignment('left')
    
    
    for i in range(12,16):
        ind1 = worksheet2.cell (row=i, column=3)
        ind1.alignment = Alignment(indent=3)
        ind2 = worksheet2.cell (row=i, column=2)
        ind2.alignment = Alignment('right')
         
    for i in range(6,12):
        ag = worksheet2.cell (row=i, column=2)
        ag.font = Font(name='Calibri', bold=True)
        ag2 = worksheet2.cell (row=i, column=3)
        ag2.font = Font(name='Calibri', bold=True)
        
    for i in range(4,7):
        ag = worksheet2.cell (row=4, column=i)
        ag.font = Font(name='Calibri', bold=True, size=12 )
        ag.alignment = Alignment(horizontal="center", vertical="center")
        
    for i in range(5,16):
        cell = worksheet2.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(5,16):
        cell = worksheet2.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
  
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    worksheet7 = workbook.create_sheet('somme Trésorerie', 6)
    
    
    for c in Canva.objects.filter(id=pk):
        tresorerieSum_queryset = Tresorerie.objects.filter(canva__mois=c.mois, canva__année=c.année)
    
    for c in Canva.objects.filter(id=pk):
        columns = ['Comptes SCF', 'intitulé', 'Mois m-1', 'Mois m']

    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet7.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for a in tresorerieSum_queryset:
        row_num += 1
        n=n+1
        row = [a.SCF, a.banques, a.moism1, a.moism]
    
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet7.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')    
            
            if cell.value is None:
                cell.value=0
                
    for i in range(5,16):
        obs = worksheet2.cell (row=i, column=6)
        if obs.value==0:
            obs.value=""
    
    
    for i in range(1, 7):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet7.column_dimensions[column_letter]
        column_dimensions.width = 25
    
    
    
    for c in range(1,5):
        for l in range (2,1880):
            cell = worksheet7.cell(row=l, column=c)
            if cell.value=="Banques, Etablissement Finaniers et Assimilés":
                for d in range(1,6):
                    ligne = worksheet7.cell(row=l, column=d)
                    ligne.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')

    
    n=2
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=5, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
        
    
    n=2
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=5, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11   

    n=3
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=6, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
    
    n=3
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            cell1.value=0
        else:
            break
        cell = worksheet2.cell(row=6, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11    

    n=4
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=7, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
    
    n=4
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=7, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11   
    
    n=5
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=8, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
    
    n=5
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            cell1.value=0
        else:
            break
        cell = worksheet2.cell(row=8, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11  

    n=6
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=9, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
    
    n=6
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=9, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11
        
    n=7
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=10, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
    
    n=7
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=10, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11    
        
    n=8
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=11, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
    
    n=8
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=11, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11 

    n=9
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=12, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
    
    n=9
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=12, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11 
        
    n=10
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=13, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
    
    n=10
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=13, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11 
        
    n=11
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=14, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
    
    n=11
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=14, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11  
        
    n=12
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=15, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+11
    
    n=12
    for i in range(1,13):
        cell1 = worksheet7.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet2.cell(row=15, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+11
        
    for i in range(5,16):
        cell = worksheet2.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(5,16):
        cell = worksheet2.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
        
     
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    worksheet3 = workbook.create_sheet('Production physique', 2)
    
    
    
    production_queryset = Production.objects.filter(canva_id=pk)
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['Production physique', 'Unité', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 6
    
    for col_num, column_title in enumerate(columns, 2):
        cell = worksheet3.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for a in production_queryset:
        row_num += 1
        n=n+1
        
        row = [a.produit, a.unité, 0, 0, 0, 0, 0, 0, 0, 0]
    
        
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet3.cell(row=row_num, column=col_num+1)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')    
            
            if cell.value is None:
                cell.value=0
        
        
            
            
    
    worksheet3.merge_cells('B1:D1')
    merged2 = worksheet3['B1']
    merged2b = worksheet3['D1']
    merged2c = worksheet3['C1']
    merged2.value = 'EPE : ANABIB Spa'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    
    
    worksheet3.merge_cells('B2:D2')
    merged2 = worksheet3['B2']
    merged2b = worksheet3['D2']
    merged2c = worksheet3['c2']
    for c in Canva.objects.filter(id=pk):
        d = c.mois      
        a = c.année
    merged2.value = 'Période :  '+d+'  '+str(a)
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='medium'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet3.merge_cells('E1:K1')
    merged2 = worksheet3['E1']
    merged2b = worksheet3['K1']
    merged2c = worksheet3['F1']
    merged2d = worksheet3['G1']
    merged2e = worksheet3['H1']
    merged2f = worksheet3['I1']
    merged2g = worksheet3['J1']
    merged2.value = 'Groupe IMETAL Spa'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='medium'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2d.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2e.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2f.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2g.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet3.merge_cells('E2:K2')
    merged2 = worksheet3['E2']
    merged2b = worksheet3['K2']
    merged2c = worksheet3['F2']
    merged2d = worksheet3['G2']
    merged2e = worksheet3['H2']
    merged2f = worksheet3['I2']
    merged2g = worksheet3['J2']
    merged2.value = 'Tbg 2 : Production physique'
    merged2.alignment = Alignment(horizontal="left", vertical="center")
    merged2.border =  Border(top=Side(border_style='medium'), left=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='medium'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
    merged2c.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2d.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2e.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2f.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2g.border =  Border(top=Side(border_style='medium'), bottom=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=13)
    
    worksheet3.merge_cells('B5:B6')
    merged2 = worksheet3['B5']
    merged2b = worksheet3['B6']
    merged2.value = "Production physique"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=15)
    
    worksheet3.merge_cells('C5:C6')
    merged2 = worksheet3['C5']
    merged2b = worksheet3['C6']
    merged2.value = "Unité"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=15)
    
    
    worksheet3.merge_cells('D5:E5')
    merged2 = worksheet3['D5']
    merged2b = worksheet3['E5']
    merged2.value = "Mois"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=12)
    
    worksheet3.merge_cells('F5:G5')
    merged2 = worksheet3['F5']
    merged2b = worksheet3['G5']
    merged2.value = "Evolution"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    
    worksheet3.merge_cells('H5:I5')
    merged2 = worksheet3['H5']
    merged2b = worksheet3['I5']
    merged2.value = "Cumul"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=12)
    
    worksheet3.merge_cells('J5:K5')
    merged2 = worksheet3['J5']
    merged2b = worksheet3['K5']
    merged2.value = "Evolution"
    merged2.alignment = Alignment(horizontal="center", vertical="center")
    merged2.border =  Border(top=Side(border_style='thick'), left=Side(border_style='thick'), bottom=Side(border_style='thick'), right=Side(border_style='thick'))
    merged2b.border =  Border(top=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'), left=Side(border_style='thick'))
    merged2.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
    merged2.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    
    cell = worksheet3.cell (row=6, column=6)
    cell.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    cell = worksheet3.cell (row=6, column=7)
    cell.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    cell = worksheet3.cell (row=6, column=10)
    cell.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    cell = worksheet3.cell (row=6, column=11)
    cell.font = Font(name='Calibri', bold=True, size=12, color="000000ff")
    
    
    
    
    
    
    
    
    
    
    
    column_letter = get_column_letter(2)
    column_dimensions = worksheet3.column_dimensions[column_letter]
    column_dimensions.width = 35
    
    column_letter = get_column_letter(3)
    column_dimensions = worksheet3.column_dimensions[column_letter]
    column_dimensions.width = 15
    
    for i in range(4,12):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet3.column_dimensions[column_letter]
        column_dimensions.width = 20
    
    for i in range(8,30):
        ag = worksheet3.cell (row=i, column=2)
        ag.alignment = Alignment('left')
        ag.font = Font(name='Calibri', bold=True)
    
    
    for i in range(10,13):
        ind1 = worksheet3.cell (row=i, column=2)
        ind1.alignment = Alignment(indent=3)
        ind1.font = Font(name='Calibri', bold=False)
    
    for i in range(7,30):
        ag = worksheet3.cell (row=i, column=3)
        ag.font = Font(name='Calibri', bold=True)
        
    for i in range(2,12):
        ag = worksheet3.cell (row=7, column=i)
        ag.font = Font(name='Calibri', bold=True)
        ag.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
    worksheet8 = workbook.create_sheet('somme prod', 8)
    
    
    for c in Canva.objects.filter(id=pk):
        sprod_queryset = Production.objects.filter(canva__mois=c.mois, canva__année=c.année)
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['Agregat', 'SCF', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet8.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for b in sprod_queryset:
        row_num += 1
        n=n+1
        row = [b.produit, b.unité, b.mois1, b.mois2]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet8.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')  
                 
            
            if cell.value is None:
                cell.value=0

    for i in range(1,5):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet8.column_dimensions[column_letter]
        column_dimensions.width = 35 
        
    for c in range(1,5):
        for l in range (2,314):
            cell = worksheet8.cell(row=l, column=c)
            if cell.value=="PRODUIT":
                for d in range(1,5):
                    ligne = worksheet8.cell(row=l, column=d)
                    ligne.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
                    
    n=2
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=7, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=2
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=7, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23 
       
    n=3
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=8, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=3
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=8, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23    

    n=4
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=9, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=4
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=9, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23   

    n=5
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=10, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=5
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=10, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23   
    
    n=6
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=11, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=6
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=11, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23  

    n=7
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=12, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=7
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=12, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23    
        
    n=8
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=13, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=8
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=13, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23    
        
    n=9
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=14, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=9
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=14, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23 

    n=10
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=15, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=10
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=15, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23 
        
    n=11
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=16, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=11
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=16, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23  
        
    n=12
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=17, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=12
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=17, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23 
        
    n=13
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=18, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=13
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=18, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=14
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=19, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=14
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=19, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=15
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=20, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=15
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=20, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=16
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=21, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=16
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=21, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=17
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=22, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=17
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=22, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23

    n=18
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=23, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=18
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=23, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=19
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=24, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=19
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=24, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=20
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=25, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=20
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=25, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=21
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=26, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=21
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=26, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=22
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=27, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=22
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=27, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=23
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=28, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=23
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=28, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23

    n=24
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=29, column=4)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=24
    for i in range(1,13):
        cell1 = worksheet8.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=29, column=5)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    
    for l in range(8,30):
        cell1 = worksheet3.cell (row=l, column=4)
        cell2 = worksheet3.cell (row=l, column=5)
        cell = worksheet3.cell (row=l, column=6)
        cell.value=cell2.value-cell1.value
        
    for l in range(8,30):
        cell1 = worksheet3.cell (row=l, column=4)
        cell2 = worksheet3.cell (row=l, column=5)
        cell = worksheet3.cell (row=l, column=7)
        if cell1.value==0:
            cell.value=100
        else:
            cell.value=(cell2.value-cell1.value)/cell1.value*100
    
    
    
    


    
        
    
  
            
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    worksheet9 = workbook.create_sheet('prodcs', 9)
    
    
    for c in Canva.objects.filter(id=pk):
        prodcs_queryset = Production.objects.filter(canva__created__year=c.année, canva__created__lte=c.created)
    
    
    for c in Canva.objects.filter(id=pk):
        columns = ['Agregat', 'SCF', c.mois+' '+'\n'+str(c.année-1), c.mois+' '+str(c.année), 'Ecart en Valeur', '% evolution', 'a la fin '+c.mois+' '+ str(c.année-1), 'a la fin '+c.mois+' '+ str(c.année), 'Ecart en Valeur', '% evolution']

    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet9.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.border = Border(top=Side(border_style='thick'), left=Side(border_style='thick'), right=Side(border_style='thick'), bottom=Side(border_style='thick'))
        cell.alignment = Alignment('center') 
        cell.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
        cell.font = Font(name='Calibri', bold=True)
        
    n = 0
    for b in prodcs_queryset:
        row_num += 1
        n=n+1
        row = [b.produit, b.unité, b.mois1, b.mois2]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet9.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = Border(bottom=Side(border_style='medium'), left=Side(border_style='thick'), top=Side(border_style='medium'), right=Side(border_style='thick'))
            cell.alignment = Alignment('center')  
                 
            
            if cell.value is None:
                cell.value=0

    for i in range(1,5):
        column_letter = get_column_letter(i)
        column_dimensions = worksheet9.column_dimensions[column_letter]
        column_dimensions.width = 35 
        
    for c in range(1,5):
        for l in range (2,1660):
            cell = worksheet9.cell(row=l, column=c)
            if cell.value=="PRODUIT":
                for d in range(1,5):
                    ligne = worksheet9.cell(row=l, column=d)
                    ligne.fill = PatternFill(start_color="0099ccff", end_color=("0099ccff"), fill_type='solid')
                    
    n=2
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=7, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=2
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=7, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23 
       
    n=3
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=8, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=3
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=8, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23    

    n=4
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=9, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=4
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=9, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23   

    n=5
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=10, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=5
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=10, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23   
    
    n=6
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=11, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=6
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=11, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23  

    n=7
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=12, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=7
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=12, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23    
        
    n=8
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=13, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=8
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=13, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23    
        
    n=9
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=14, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=9
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=14, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23 

    n=10
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=15, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=10
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=15, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23 
        
    n=11
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=16, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=11
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=16, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23  
        
    n=12
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=17, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=12
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=17, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23 
        
    n=13
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=18, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=13
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=18, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=14
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=19, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=14
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=19, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=15
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=20, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=15
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=20, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=16
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=21, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=16
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=21, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=17
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=22, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=17
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=22, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23

    n=18
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=23, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=18
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=23, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=19
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=24, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=19
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=24, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=20
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=25, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=20
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=25, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=21
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=26, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=21
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=26, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=22
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=27, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=22
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=27, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
        
    n=23
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=28, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=23
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=28, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23

    n=24
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=3)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=29, column=8)
        cell.value = int(cell.value+cell1v)
        n=n+23
    
    n=24
    for i in range(1,75):
        cell1 = worksheet9.cell(row=n, column=4)
        if cell1.value is None:
            break
        else:
            cell1v = cell1.value
        cell = worksheet3.cell(row=29, column=9)
        cell.value = int(cell.value+cell1v)
        n=n+23
                    
                    
    for l in range(8,30):
        cell1 = worksheet3.cell (row=l, column=8)
        cell2 = worksheet3.cell (row=l, column=9)
        cell = worksheet3.cell (row=l, column=10)
        cell.value=cell2.value-cell1.value
        
    for l in range(8,30):
        cell1 = worksheet3.cell (row=l, column=8)
        cell2 = worksheet3.cell (row=l, column=9)
        cell = worksheet3.cell (row=l, column=11)
        if cell1.value==0:
            cell.value=100
        else:
            cell.value=(cell2.value-cell1.value)/cell1.value*100
            
            
    
        
    
    
    
    for i in range(8,34):
        cell = worksheet3.cell (row=i, column=4)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(8,34):
        cell = worksheet3.cell (row=i, column=5)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(8,30):
        cell = worksheet3.cell (row=i, column=6)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(8,30):
        cell = worksheet3.cell (row=i, column=8)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(8,30):
        cell = worksheet3.cell (row=i, column=9)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    for i in range(8,30):
        cell = worksheet3.cell (row=i, column=10)
        if cell.value!=0:
            cell.number_format = '### ### ### ###'
            
    
    
            
    for i in range(8,30):   
        cell = worksheet3.cell (row=i, column=11)
        cell.number_format = '0.00'
        
    for i in range(8,30):   
        cell = worksheet3.cell (row=i, column=7)
        cell.number_format = '0.00'
    
    for c in range(4,6):
        cell = worksheet2.cell (row=5, column=c)
        cell.value=""
        
    for c in range(4,12):
        cell = worksheet3.cell (row=7, column=c)
        cell.value=""

    
    workbook.save(response) 
    return response
        
        
       
        
            
        


    

    